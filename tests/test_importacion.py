import pytest
from openpyxl import load_workbook

from app.db.init_db import init_database
from app.db.seed import sembrar_configuracion
from app.importacion.importar_excel import importar_planilla
from app.importacion.plantillas import generar_plantillas
from app.repositorio.registro import obtener_repositorio


@pytest.fixture
def conn(tmp_path):
    connection = init_database(tmp_path / "test.db")
    yield connection
    connection.close()


def test_generar_plantilla_crea_una_hoja_por_entidad(tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    assert ruta.exists()

    wb = load_workbook(ruta)
    assert "Edificio" in wb.sheetnames
    assert "Unidad" in wb.sheetnames
    assert "Consultorio" in wb.sheetnames
    assert wb["Edificio"][1][0].value == "Nombre"


def test_importar_edificio_unidad_consultorio(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)

    wb["Edificio"].append(["Ramos 1", "Av. Rivadavia 13876", "Ramos Mejía"])

    wb["Unidad"].append(
        ["Ramos 1", '7mo "L"', "SI", "SI", 2, "NO", "NO", "NO", "NO", "NO", "NO", "SI", 60]
    )

    wb["Consultorio"].append(
        ["Ramos 1", '7mo "L"', 3, 4.2, 3.5, "intermedio", "SI", "NO", "SI", "NO", "SI", "SI", "NO", 4646, 4646]
    )

    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}

    assert resultados["Edificio"].filas_importadas == 1
    assert resultados["Unidad"].filas_importadas == 1
    assert resultados["Consultorio"].filas_importadas == 1
    assert not resultados["Unidad"].errores
    assert not resultados["Consultorio"].errores

    edificio = obtener_repositorio(conn, "Edificio").listar()[0]
    unidad = obtener_repositorio(conn, "Unidad").listar()[0]
    consultorio = obtener_repositorio(conn, "Consultorio").listar()[0]

    assert unidad["IdEdificio"] == edificio["IdEdificio"]
    assert unidad["WiFi"] == 1
    assert consultorio["IdUnidad"] == unidad["IdUnidad"]
    assert consultorio["NumeroConsultorio"] == 3


def test_importar_referencia_inexistente_reporta_error(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)

    wb["Unidad"].append(
        ["Edificio Que No Existe", "PB", "NO", "NO", 1, "NO", "NO", "NO", "NO", "NO", "NO", "NO", 10]
    )
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}

    assert resultados["Unidad"].filas_importadas == 0
    assert len(resultados["Unidad"].errores) == 1
    assert "Edificio Que No Existe" in resultados["Unidad"].errores[0]


def test_importar_convierte_fecha_dd_mm_aaaa_a_iso(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    wb["Edificio"].append(["Ramos 1", "Av. Rivadavia 13876", "Ramos Mejía"])
    wb["Unidad"].append(
        ["Ramos 1", '7mo "L"', "SI", "SI", 2, "NO", "NO", "NO", "NO", "NO", "NO", "SI", 60]
    )
    wb["Consultorio"].append(
        ["Ramos 1", '7mo "L"', 3, 4.2, 3.5, "intermedio", "SI", "NO", "SI", "NO", "SI", "SI", "NO", 4646, 4646]
    )
    wb["Profesional"].append(
        ["R", "R1", "Lo Veci", "María Virginia Lo Veci", "Virginia", "Virgi", "Femenino",
         "28456789", "27-28456789-3", "Monotributo", "15/03/1985", "Av. San Martín 1234", "Ramos Mejía",
         None, "1145678901", None, None, "Lic.", None, None, None, None, None, None, None]
    )
    wb["ReservaRegular"].append(
        ["R1", "Ramos 1", '7mo "L"', "3", "Lunes", "14", "18", "01/08/2026", None, "NO", None]
    )
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert not resultados["Profesional"].errores
    assert not resultados["ReservaRegular"].errores

    profesional = obtener_repositorio(conn, "Profesional").listar()[0]
    assert profesional["FechaNacimiento"] == "1985-03-15"
    reserva = obtener_repositorio(conn, "ReservaRegular").listar()[0]
    assert reserva["VigenciaInicio"] == "2026-08-01"


def test_importar_fecha_invalida_reporta_error_de_fila(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    wb["Edificio"].append(["Ramos 1", "Av. Rivadavia 13876", "Ramos Mejía"])
    wb["Unidad"].append(
        ["Ramos 1", '7mo "L"', "SI", "SI", 2, "NO", "NO", "NO", "NO", "NO", "NO", "SI", 60]
    )
    wb["Consultorio"].append(
        ["Ramos 1", '7mo "L"', 3, 4.2, 3.5, "intermedio", "SI", "NO", "SI", "NO", "SI", "SI", "NO", 4646, 4646]
    )
    wb["Profesional"].append(
        ["R", "R1", "Lo Veci", "María Virginia Lo Veci", "Virginia", "Virgi", "Femenino",
         "28456789", "27-28456789-3", "Monotributo", "1985-03-15", "Av. San Martín 1234", "Ramos Mejía",
         None, "1145678901", None, None, "Lic.", None, None, None, None, None, None, None]
    )
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert resultados["Profesional"].filas_importadas == 0
    assert len(resultados["Profesional"].errores) == 1
    assert "FechaNacimiento" in resultados["Profesional"].errores[0]


def test_importar_profesional_cabeza_de_equipo(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    fila_r = ["R", "R1", "Lo Veci", "María Virginia Lo Veci", "Virginia", "Virgi", "Femenino",
              "28456789", "27-28456789-3", "Monotributo", None, None, None, None, "1145678901",
              None, None, "Lic.", None, None, None, None, None, None, None, None]
    fila_e = ["E", "E1", "Gomez", "Ana Gomez", "Ana", None, "Femenino",
              "30111222", "27-30111222-3", "Monotributo", None, None, None, None, "1145678902",
              None, None, "Lic.", None, None, None, None, "R1", None, None, None]
    wb["Profesional"].append(fila_r)
    wb["Profesional"].append(fila_e)
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert not resultados["Profesional"].errores
    assert resultados["Profesional"].filas_importadas == 2

    r1 = obtener_repositorio(conn, "Profesional").listar(IdCodigo="R1")[0]
    e1 = obtener_repositorio(conn, "Profesional").listar(IdCodigo="E1")[0]
    assert e1["ProfesionalCabezaEquipo"] == r1["IdProfesional"]


def test_importar_saldo_cuenta_anterior(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    encabezados = [c.value for c in wb["Profesional"][1]]
    assert "SaldoCuentaAnterior" in encabezados
    columna_saldo = encabezados.index("SaldoCuentaAnterior")

    fila = [None] * len(encabezados)
    fila[encabezados.index("CategoriaProfesional")] = "R"
    fila[encabezados.index("IdCodigo")] = "R1"
    fila[encabezados.index("Apellido")] = "Lo Veci"
    fila[columna_saldo] = 15000.50
    wb["Profesional"].append(fila)
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert not resultados["Profesional"].errores

    r1 = obtener_repositorio(conn, "Profesional").listar(IdCodigo="R1")[0]
    assert r1["SaldoCuentaAnterior"] == 15000.50


def test_importar_fechas_especiales(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    wb["FechasEspeciales"].append(["25/05/2026", "Día de la Patria", "Feriado nacional"])
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert resultados["FechasEspeciales"].filas_importadas == 1
    fecha = obtener_repositorio(conn, "FechasEspeciales").listar()[0]
    assert fecha["Fecha"] == "2026-05-25"
    assert fecha["Tipo"] == "Feriado nacional"


# --------------------------------------------------------------- PlanPago

def _fijar_fecha(conn, fecha_iso: str) -> None:
    sembrar_configuracion(conn)
    conn.execute(
        "UPDATE Configuracion SET ModoFechaFicticia = 1, FechaFicticia = ? WHERE IdConfiguracion = 1",
        (fecha_iso,),
    )
    conn.commit()


def _agregar_profesional_r1(wb) -> None:
    encabezados = [c.value for c in wb["Profesional"][1]]
    fila = [None] * len(encabezados)
    fila[encabezados.index("CategoriaProfesional")] = "R"
    fila[encabezados.index("IdCodigo")] = "R1"
    fila[encabezados.index("Apellido")] = "Lo Veci"
    wb["Profesional"].append(fila)


def test_importar_plan_pago_calcula_cuota_y_marca_pasadas_pagadas(conn, tmp_path):
    """Plan acordado en mayo por 12 cuotas de $1000, importado el 15/8: las
    cuotas de mayo/junio/julio (ya transcurridas) quedan solas como
    pagadas, la de agosto (período actual) y las siguientes, pendientes."""
    _fijar_fecha(conn, "2026-08-15")
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    _agregar_profesional_r1(wb)
    wb["PlanPago"].append(["R1", "05/2026", 12000, 0, 12, "Plan acordado antes del sistema"])
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert not resultados["PlanPago"].errores
    assert resultados["PlanPago"].filas_importadas == 1

    plan = obtener_repositorio(conn, "PlanPago").listar()[0]
    assert plan["MesAnoInicio"] == "2026-05"
    assert plan["CantidadCuotas"] == 12
    assert plan["ImportePorCuota"] == 1000.0
    assert plan["Estado"] == "Activo"

    cuotas = {c["PeriodoImputado"]: c for c in obtener_repositorio(conn, "CuotaPlan").listar(IdPlan=plan["IdPlan"])}
    assert cuotas["2026-05"]["Estado"] == "Pagada"
    assert cuotas["2026-06"]["Estado"] == "Pagada"
    assert cuotas["2026-07"]["Estado"] == "Pagada"
    assert cuotas["2026-08"]["Estado"] == "Pendiente"
    assert cuotas["2026-09"]["Estado"] == "Pendiente"


def test_importar_plan_pago_calcula_interes(conn, tmp_path):
    _fijar_fecha(conn, "2026-08-15")
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    _agregar_profesional_r1(wb)
    wb["PlanPago"].append(["R1", "08/2026", 10000, 5, 4, None])
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert not resultados["PlanPago"].errores
    plan = obtener_repositorio(conn, "PlanPago").listar()[0]
    # 10000 * (1 + 0.05 * 4) = 12000 total / 4 cuotas = 3000 cada una.
    assert plan["MontoTotalAPagar"] == 12000.0
    assert plan["ImportePorCuota"] == 3000.0


def test_importar_plan_pago_referencia_profesional_inexistente(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    wb["PlanPago"].append(["R99", "05/2026", 12000, 0, 12, None])
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert resultados["PlanPago"].filas_importadas == 0
    assert "R99" in resultados["PlanPago"].errores[0]


def test_importar_plan_pago_periodo_invalido_reporta_error(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    _agregar_profesional_r1(wb)
    wb["PlanPago"].append(["R1", "2026-05", 12000, 0, 12, None])  # formato incorrecto
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert resultados["PlanPago"].filas_importadas == 0
    assert "MesAnoInicio" in resultados["PlanPago"].errores[0]


def test_importar_plan_pago_falta_columna_obligatoria(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    _agregar_profesional_r1(wb)
    wb["PlanPago"].append(["R1", "05/2026", None, 0, 12, None])  # sin MontoRefinanciado
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert resultados["PlanPago"].filas_importadas == 0
    assert "MontoRefinanciado" in resultados["PlanPago"].errores[0]


def test_importar_plan_pago_ya_existe_en_plantilla(tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    assert "PlanPago" in wb.sheetnames
    assert [c.value for c in wb["PlanPago"][1]] == [
        "Profesional", "MesAnoInicio", "MontoRefinanciado",
        "PorcentajeInteresMensual", "CantidadCuotas", "Observacion",
    ]


def test_importar_alias_banos_con_tilde(conn, tmp_path):
    ruta = generar_plantillas(tmp_path / "plantilla.xlsx")
    wb = load_workbook(ruta)
    wb["Edificio"].append(["Ramos 1", "Av. Rivadavia 13876", "Ramos Mejía"])
    ws_unidad = wb["Unidad"]
    ws_unidad.cell(row=1, column=5, value="Baños")  # encabezado con tilde, en vez de "Banos"
    ws_unidad.append(["Ramos 1", '7mo "L"', "SI", "SI", 2, "NO", "NO", "NO", "NO", "NO", "NO", "SI", 60])
    wb.save(ruta)

    resultados = {r.entidad: r for r in importar_planilla(conn, ruta)}
    assert not resultados["Unidad"].errores
    unidad = obtener_repositorio(conn, "Unidad").listar()[0]
    assert unidad["Banos"] == 2
