import pytest
from PySide6.QtWidgets import QMessageBox

from app.db.init_db import init_database
from app.db.seed import sembrar_valores_por_defecto
from app.gui.pantallas.importacion import PantallaImportacion


@pytest.fixture
def conn(tmp_path):
    connection = init_database(tmp_path / "test.db")
    sembrar_valores_por_defecto(connection)
    yield connection
    connection.close()


@pytest.fixture(autouse=True)
def _sin_dialogos_modales(monkeypatch):
    monkeypatch.setattr(QMessageBox, "information", staticmethod(lambda *a, **k: None))
    monkeypatch.setattr(QMessageBox, "critical", staticmethod(lambda *a, **k: None))


def _planilla_minima(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Edificio")
    ws.append(["Nombre", "Domicilio", "DomicilioLocalidad"])
    ws.append(["Torre Norte", "Calle 1", "Ramos Mejía"])
    ruta = tmp_path / "planilla.xlsx"
    wb.save(ruta)
    return str(ruta)


def test_boton_importar_deshabilitado_sin_archivo(qtbot, conn):
    pantalla = PantallaImportacion(conn)
    qtbot.addWidget(pantalla)
    assert pantalla.boton_importar.isEnabled() is False


def test_importar_carga_datos_y_llena_tabla_resultados(qtbot, conn, tmp_path):
    pantalla = PantallaImportacion(conn)
    qtbot.addWidget(pantalla)
    pantalla.campo_ruta.setText(_planilla_minima(tmp_path))
    pantalla._importar()

    assert pantalla.tabla_resultados.rowCount() == 1
    assert pantalla.tabla_resultados.item(0, 0).text() == "Edificio"
    assert pantalla.tabla_resultados.item(0, 1).text() == "1"
    assert conn.execute("SELECT COUNT(*) c FROM Edificio").fetchone()["c"] == 1


def test_importar_sin_problemas_muestra_sin_observaciones(qtbot, conn, tmp_path):
    pantalla = PantallaImportacion(conn)
    qtbot.addWidget(pantalla)
    pantalla.campo_ruta.setText(_planilla_minima(tmp_path))
    pantalla._importar()
    assert pantalla.texto_integridad.toPlainText() == "Sin observaciones."


def test_informe_muestra_codigo_duplicado(qtbot, conn, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Profesional")
    ws.append(["CategoriaProfesional", "IdCodigo", "Apellido"])
    ws.append(["R", "R1", "Perez"])
    ws.append(["R", "R1", "Gomez"])
    ruta = tmp_path / "planilla.xlsx"
    wb.save(ruta)

    pantalla = PantallaImportacion(conn)
    qtbot.addWidget(pantalla)
    pantalla.campo_ruta.setText(str(ruta))
    pantalla._importar()

    assert "R1" in pantalla.texto_integridad.toPlainText()
    assert "duplicado" in pantalla.texto_integridad.toPlainText()
