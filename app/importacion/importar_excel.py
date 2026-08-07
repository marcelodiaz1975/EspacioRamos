"""Importación de datos iniciales desde la planilla Excel (FA5).

Las columnas de referencia (Edificio, Unidad, Profesional, etc.) se cargan
en la planilla con nombres legibles en vez de IDs internos -- por ejemplo,
en la hoja "Consultorio" se indica el nombre del Edificio y el Departamento
de la Unidad, no sus IdEdificio/IdUnidad. Este módulo resuelve esas
referencias contra lo que ya se cargó en la base (por eso el orden de
ENTIDADES_IMPORTABLES importa: primero Edificio, después Unidad, etc.)

Fechas (DC-11 caso 7): el único formato aceptado en la planilla es
DD/MM/AAAA; se convierten a ISO (AAAA-MM-DD) antes de guardar, y un
formato inválido se reporta como error de esa fila/columna sin abortar el
resto de la importación. openpyxl también puede entregar la celda ya como
datetime si Excel la formateó como fecha -- se acepta igual.
"""
from __future__ import annotations

import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.importacion.definiciones import (
    ALIAS_ENCABEZADOS,
    CAMPOS_BOOLEANOS,
    CAMPOS_FECHA,
    COLUMNAS_REFERENCIA,
    ENTIDADES_IMPORTABLES,
)
from app.repositorio.registro import obtener_repositorio


@dataclass
class ResultadoImportacion:
    entidad: str
    filas_importadas: int = 0
    errores: list[str] = field(default_factory=list)


def _convertir_booleano(valor) -> int:
    texto = str(valor).strip().upper()
    return 1 if texto in ("SI", "SÍ", "TRUE", "1", "X") else 0


def _convertir_fecha(valor) -> str:
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    texto = str(valor).strip()
    partes = texto.split("/")
    if len(partes) != 3:
        raise ValueError(f"Fecha inválida (se espera DD/MM/AAAA): {valor!r}")
    dia, mes, anio = partes
    try:
        return date(int(anio), int(mes), int(dia)).isoformat()
    except ValueError as exc:
        raise ValueError(f"Fecha inválida (se espera DD/MM/AAAA): {valor!r}") from exc


def _convertir_valor(columna: str, valor):
    if valor is None or valor == "":
        return None
    if columna in CAMPOS_BOOLEANOS:
        return _convertir_booleano(valor)
    if columna in CAMPOS_FECHA:
        return _convertir_fecha(valor)
    return valor


def _buscar_id(conn: sqlite3.Connection, tabla: str, columna: str, valor) -> int | None:
    if valor is None:
        return None
    fila = conn.execute(f"SELECT * FROM {tabla} WHERE {columna} = ? LIMIT 1", (valor,)).fetchone()
    return fila[0] if fila else None


def _buscar_unidad(conn: sqlite3.Connection, nombre_edificio, departamento) -> int | None:
    id_edificio = _buscar_id(conn, "Edificio", "Nombre", nombre_edificio)
    if id_edificio is None:
        return None
    fila = conn.execute(
        "SELECT IdUnidad FROM Unidad WHERE IdEdificio = ? AND Departamento = ?",
        (id_edificio, departamento),
    ).fetchone()
    return fila["IdUnidad"] if fila else None


def _buscar_consultorio(conn: sqlite3.Connection, nombre_edificio, departamento, numero) -> int | None:
    id_unidad = _buscar_unidad(conn, nombre_edificio, departamento)
    if id_unidad is None:
        return None
    fila = conn.execute(
        "SELECT IdConsultorio FROM Consultorio WHERE IdUnidad = ? AND NumeroConsultorio = ?",
        (id_unidad, numero),
    ).fetchone()
    return fila["IdConsultorio"] if fila else None


def _resolver_referencias(conn: sqlite3.Connection, entidad: str, datos: dict) -> tuple[dict, list[str]]:
    """Reemplaza las columnas de referencia (nombres legibles) por su IdX.
    Devuelve los datos resueltos y una lista de errores (referencias no
    encontradas)."""
    errores: list[str] = []
    referencias = COLUMNAS_REFERENCIA.get(entidad, set())
    if not referencias:
        return datos, errores

    if entidad == "Unidad":
        nombre_ed = datos.pop("Edificio", None)
        id_edificio = _buscar_id(conn, "Edificio", "Nombre", nombre_ed)
        if id_edificio is None:
            errores.append(f"No se encontró el edificio '{nombre_ed}'")
        datos["IdEdificio"] = id_edificio

    elif entidad == "Consultorio":
        nombre_ed = datos.pop("Edificio", None)
        depto = datos.pop("Unidad", None)
        id_unidad = _buscar_unidad(conn, nombre_ed, depto)
        if id_unidad is None:
            errores.append(f"No se encontró la unidad '{depto}' del edificio '{nombre_ed}'")
        datos["IdUnidad"] = id_unidad

    elif entidad == "Profesional":
        nombre_prof = datos.pop("Profesion", None)
        if nombre_prof is not None:
            id_profesion = _buscar_id(conn, "Profesion", "Nombre", nombre_prof)
            if id_profesion is None:
                errores.append(f"No se encontró la profesión '{nombre_prof}'")
            datos["IdProfesion"] = id_profesion

        codigo_cabeza = datos.pop("ProfesionalCabezaEquipo", None)
        if codigo_cabeza is not None:
            id_cabeza = _buscar_id(conn, "Profesional", "IdCodigo", codigo_cabeza)
            if id_cabeza is None:
                errores.append(
                    f"No se encontró el profesional cabeza de equipo con código '{codigo_cabeza}' "
                    "(tiene que estar cargado antes en la planilla)"
                )
            datos["ProfesionalCabezaEquipo"] = id_cabeza

    elif entidad == "ReservaRegular":
        codigo = datos.pop("Profesional", None)
        id_profesional = _buscar_id(conn, "Profesional", "IdCodigo", codigo)
        if id_profesional is None:
            errores.append(f"No se encontró el profesional con código '{codigo}'")
        datos["IdProfesional"] = id_profesional

        nombre_ed = datos.pop("Edificio", None)
        depto = datos.pop("Unidad", None)
        numero = datos.pop("Consultorio", None)
        id_consultorio = _buscar_consultorio(conn, nombre_ed, depto, numero)
        if id_consultorio is None:
            errores.append(
                f"No se encontró el consultorio {numero} de la unidad '{depto}' del edificio '{nombre_ed}'"
            )
        datos["IdConsultorio"] = id_consultorio

    elif entidad == "Placa":
        nombre_ed = datos.pop("Edificio", None)
        depto = datos.pop("Unidad", None)
        id_unidad = _buscar_unidad(conn, nombre_ed, depto)
        if id_unidad is None:
            errores.append(f"No se encontró la unidad '{depto}' del edificio '{nombre_ed}'")
        datos["IdUnidad"] = id_unidad

        codigo = datos.pop("Profesional", None)
        if codigo is not None:
            id_profesional = _buscar_id(conn, "Profesional", "IdCodigo", codigo)
            if id_profesional is None:
                errores.append(f"No se encontró el profesional con código '{codigo}'")
            datos["IdProfesional"] = id_profesional

    return datos, errores


def importar_hoja(conn: sqlite3.Connection, entidad: str, ws) -> ResultadoImportacion:
    resultado = ResultadoImportacion(entidad=entidad)
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return resultado

    encabezados = [
        ALIAS_ENCABEZADOS.get(str(c).strip(), str(c).strip()) if c is not None else "" for c in filas[0]
    ]
    repo = obtener_repositorio(conn, entidad)

    for numero_fila, fila in enumerate(filas[1:], start=2):
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue

        datos: dict = {}
        error_conversion = None
        for columna, valor in zip(encabezados, fila):
            if not columna:
                continue
            try:
                datos[columna] = _convertir_valor(columna, valor)
            except ValueError as exc:
                error_conversion = f"Columna '{columna}': {exc}"
                break
        if error_conversion:
            resultado.errores.append(f"Fila {numero_fila}: {error_conversion}")
            continue

        datos, errores_referencia = _resolver_referencias(conn, entidad, datos)
        if errores_referencia:
            for err in errores_referencia:
                resultado.errores.append(f"Fila {numero_fila}: {err}")
            continue

        try:
            repo.crear(**datos)
            resultado.filas_importadas += 1
        except sqlite3.Error as exc:
            resultado.errores.append(f"Fila {numero_fila}: {exc}")

    return resultado


def importar_planilla(conn: sqlite3.Connection, ruta_excel: Path | str) -> list[ResultadoImportacion]:
    wb = load_workbook(ruta_excel, data_only=True)
    resultados = []
    for entidad in ENTIDADES_IMPORTABLES:
        if entidad not in wb.sheetnames:
            continue
        resultados.append(importar_hoja(conn, entidad, wb[entidad]))
    return resultados
