"""Generación de la planilla Excel modelo para la carga inicial de datos (FA5)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.importacion.definiciones import COLUMNAS_PLANTILLA

TITULO_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
TITULO_FONT = Font(color="FFFFFF", bold=True)


def _armar_hoja(ws: Worksheet, columnas: list[str]) -> None:
    ws.append(columnas)
    for celda in ws[1]:
        celda.fill = TITULO_FILL
        celda.font = TITULO_FONT
    for i, nombre in enumerate(columnas, start=1):
        ancho = max(14, len(nombre) + 2)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho


def _armar_hoja_instrucciones(ws: Worksheet) -> None:
    ws.append(["Instrucciones para completar la planilla"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["- Completar los datos a partir de la fila 2 de cada hoja (no dejar filas vacías en el medio)."])
    ws.append(["- Columnas de SI / NO: escribir SI o NO."])
    ws.append(["- Columnas que nombran otra entidad (por ejemplo \"Edificio\" en la hoja Unidad,"])
    ws.append(["  o \"Profesional\" en la hoja ReservaRegular) se completan con el nombre o código"])
    ws.append(["  ya cargado en la hoja correspondiente, no con un número interno."])
    ws.append(["- En la hoja ReservaRegular, Placa y PlanPago, la columna \"Profesional\" se completa"])
    ws.append(["  con el Código del profesional (columna IdCodigo de la hoja Profesional)."])
    ws.append(["- Columnas de fecha (FechaNacimiento, VigenciaInicio, VigenciaFin, Fecha, etc.):"])
    ws.append(["  formato DD/MM/AAAA."])
    ws.append(["- ProfesionalCabezaEquipo (solo para categoría E) se completa con el Código del R,"])
    ws.append(["  y ese R tiene que estar cargado ANTES en la misma hoja."])
    ws.append([])
    ws.append(["Hoja PlanPago (opcional): para dar de alta planes de pago ya acordados y en curso"])
    ws.append(["antes de empezar a usar el sistema (no hace falta para planes nuevos, esos se cargan"])
    ws.append(["desde la pantalla de Planes de pago)."])
    ws.append(["- MesAnoInicio: formato MM/AAAA (mes de la primera cuota, no el de hoy)."])
    ws.append(["- MontoRefinanciado: el monto original acordado, sin el interés."])
    ws.append(["- PorcentajeInteresMensual: dejar vacío o en 0 si no hay interés."])
    ws.append(["- CantidadCuotas: la cantidad total de cuotas del plan (no las que faltan)."])
    ws.append(["- El importe de cada cuota lo calcula el sistema solo (MontoRefinanciado más el interés,"])
    ws.append(["  dividido CantidadCuotas, todas las cuotas iguales)."])
    ws.append(["- Las cuotas de meses anteriores al de hoy se marcan solas como ya pagadas; no hace"])
    ws.append(["  falta indicar en qué cuota está el profesional actualmente."])
    ws.append([])
    ws.append(["Orden recomendado de carga: Edificio, Unidad, Consultorio, Profesion, Profesional,"])
    ws.append(["ReservaRegular, Llave, Placa, FechasEspeciales, Responsable, PlanPago."])
    ws.column_dimensions["A"].width = 100


def generar_plantillas(destino: Path | str) -> Path:
    """Genera un único libro Excel con una hoja de instrucciones y una hoja
    por entidad importable (ver app.importacion.definiciones.COLUMNAS_PLANTILLA)."""
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws_instrucciones = wb.create_sheet(title="Instrucciones")
    _armar_hoja_instrucciones(ws_instrucciones)

    for entidad, columnas in COLUMNAS_PLANTILLA.items():
        ws = wb.create_sheet(title=entidad[:31])
        _armar_hoja(ws, columnas)
    wb.save(destino)
    return destino
